import pandas as pd
from openpyxl import Workbook, styles
from openpyxl.utils.dataframe import dataframe_to_rows
table1 = pd.DataFrame({
    'A-GRAF': ['', '', ''],
    '': ['', '', ''],
   'Name': ['John', 'Alice', 'Bob'],
   'Age': [25, 30, 35],
   'City': ['New York', 'Los Angeles', 'Chicago']
})

table2 = pd.DataFrame({
   'Product': ['iPhone', 'Samsung', 'Google Pixel'],
   'Price': [1000, 800, 700],
   'Quantity': [10, 20, 30]
})

workbook = Workbook()
worksheet = workbook.active # Select the active worksheet
worksheet.title = 'Data Tables' # Rename the worksheet if needed
# Append Table 1 to cell A1
for r in dataframe_to_rows(table1, index=False, header=True):
    worksheet.append(r)

# Append Table 2 to cell A7

worksheet.cell(row=9, column=2).value = 'Table 2'
worksheet.cell(row=9, column=2).font = styles.Font(color='FF0000', size=16, bold=True, name='Calibri')
worksheet.cell(row=9, column=2).fill =styles.PatternFill(bgColor="FFC7CE", fill_type="solid")
for r in dataframe_to_rows(table2, index=False, header=True):
    worksheet.append(r)
workbook.save(filename='output.xlsx')
