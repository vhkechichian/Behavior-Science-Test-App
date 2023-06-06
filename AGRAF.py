from PyQt6.QtWidgets import QApplication,QSplashScreen,QMessageBox, QCheckBox,QAbstractItemView,QHeaderView, QFileDialog, QRadioButton,QTableWidget,QTableWidgetItem, QDialog,QSpinBox, QFormLayout, QLineEdit, QComboBox, QLabel, QMainWindow, QPushButton, QVBoxLayout, QHBoxLayout, QWidget, QGroupBox
from PyQt6.QtGui import QFont, QIcon, QPixmap, QKeyEvent, QMouseEvent, QColor
from PyQt6.QtCore import QSize,Qt, QCoreApplication, QTimer, pyqtSignal
import sys
import random
import string
import res_rcc
import time
import threading
from openpyxl import Workbook, styles

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("A-GRAF Session Launcher")
        self.setStyleSheet("QMainWindow{background-color: white} ")
        self.setFixedSize(QSize(400,270))
        self.window_new_participant = ParticipantWindow(self)
        self.window_configuration = ConfigWindow(self)

        self.btn_new = QPushButton("New Participant")
        self.btn_new.setStyleSheet("QPushButton{border: 2px solid;border-radius: 5px; background-color: white; color:black;} QPushButton:hover{background-color:grey;}")
        self.btn_new.setFixedSize(150,50)
        self.btn_new.setIconSize(QSize(40,40))
        self.btn_new.setToolTip("New Participant")
        self.btn_new.clicked.connect(self.show_new_participant)

        self.btn_config = QPushButton("Configuration")
        self.btn_config.setStyleSheet("QPushButton{border: 2px solid;border-radius: 5px; background-color: white; color:black;} QPushButton:hover{background-color:grey;}")
        self.btn_config.setFixedSize(150,50)
        self.btn_config.setIconSize(QSize(40,40))
        self.btn_config.setToolTip("Configuration")
        self.btn_config.clicked.connect(self.show_configuration)

        # btn_practice = QPushButton()
        # btn_practice.setFixedSize(100,100)
        # btn_practice.setIconSize(QSize(80,80))
        # btn_practice.setToolTip("Start Practice Session")
        # btn_practice.clicked.connect(lambda: self.show_session_window(0))

        btn_test = QPushButton()
        btn_test.setStyleSheet("QPushButton{border: 2px solid;border-radius: 5px; background-color: white; color:black;} QPushButton:hover{background-color:grey;}")
        btn_test.setIcon(QIcon(":test.jpg"))
        btn_test.setFixedSize(100,100)
        btn_test.setIconSize(QSize(80,80))
        btn_test.setToolTip("Start Test Session")
        btn_test.clicked.connect(lambda: self.show_session_window(0))

        btn_exit = QPushButton("Exit")
        btn_exit.setStyleSheet("QPushButton{border: 2px solid;border-radius: 5px; background-color: white; color:black;} QPushButton:hover{background-color:grey;}")
        # btn_exit.setIcon(QIcon(":exit.jpg"))
        btn_exit.setFixedSize(150,50)
        btn_exit.setIconSize(QSize(40,40))
        btn_exit.setToolTip("Exit")
        btn_exit.clicked.connect(self.close)
        lay_user = QHBoxLayout()
        self.btn_user = QPushButton()
        self.btn_user.setFlat(True)
        self.btn_user.setFont(QFont("Arial", 12, -1, False))
        self.btn_user.setFixedSize(100,50)
        self.btn_user.clicked.connect(self.clearUserInfo)

        main_lay = QHBoxLayout()
        top_lay = QVBoxLayout()
        center_lay = QHBoxLayout()
        top_lay.addWidget(self.btn_new)
        top_lay.addWidget(self.btn_config)
        top_lay.addStretch()
        self.lbl_user = QLabel("User ID:")
        self.lbl_user.setFont(QFont("Arial",12,-1, False))
        lay_user.addWidget(self.lbl_user)
        lay_user.addWidget(self.btn_user)
        top_lay.addLayout(lay_user)
        top_lay.addWidget(btn_exit)
        # center_lay.addWidget(btn_practice)
        center_lay.addWidget(btn_test)
        main_lay.addLayout(top_lay)
        
        main_group = QGroupBox()
        main_group.setLayout(center_lay)
        main_lay.addWidget(main_group)
        main_widget = QWidget()
        main_widget.setLayout(main_lay)
        self.setCentralWidget(main_widget)
        self.clearUserInfo()

    def show_new_participant(self):
        self.window_new_participant.exec()

    def show_session_window(self, type):
        if self.btn_user.text() == "":
            QMessageBox.warning(self,"Warning!", "There isn't participant data!")
        else:
            window_practice_session = SessionWindow(self, type)
            window_practice_session.showFullScreen()
        
    def show_configuration(self):
        self.window_configuration.exec()

    def onClickedPractice(self):
        self.btn_user.setText(self.userId)

    def clearUserInfo(self):
        self.userId = ""
        self.userAge = 0
        self.userGender = ""
        self.userHandedness = ""
        self.userEducation = ""
        self.userRace = ""
        self.userLearning = ""
        self.btn_user.setText("")
class PerTrialData():
    def __init__(self, trialNum) -> None:
        self.trialNumber = trialNum
        self.delayCount = 0
        self.shortCount = 0
        self.switches = 0
        self.trialLength = 0
        self.initialSelectionTime = 0
        self.secondSelectionTime = 0
        self.wonOn = 2
        self.firstSelection = 2
        self.payoff = 0
        self.total = 0
        self.deliveryClicks = 0
        self.itiClicks = 0
        self.nonObjectClicks = 0
        self.firstPosition = 0
    def getSecondSelectionTime(self):
        if self.secondSelectionTime > self.initialSelectionTime:
            return self.secondSelectionTime - self.initialSelectionTime
        else:
            return 0
    def getTotalCount(self):
        return self.delayCount + self.shortCount + self.nonObjectClicks
    def delayPayoff(self):
        if self.wonOn == 1:
            return self.payoff
        else:
            return 0
    def shortPayoff(self):
        if self.wonOn == 0:
            return self.payoff
        else:
            return 0
class DataLog():
    def __init__(self, numberOfTrials):
        self.preTestCount = 0
        self.shortCount = 0
        self.delayCount = 0
        self.shortConsecutive = 0
        self.currentShortStreak = 0
        self.delayConsecutive = 0
        self.currentDelayStreak = 0
        self.bankTotal = 0
        self.lastClickedObject = 2
        self.TRIAL_NUM = "Trial #"
        self.PRETEST = "Pre-test clicks"
        self.DELAY_CLICKS = "Delay clicks"
        self.SHORT_CLICKS = "Short clicks"
        self.TOTAL_CLICKS = "Total clicks"
        self.SWITCHES = "# Switches"
        self.TRIAL_LENGTH = "Trial seconds"
        self.INITIAL_SELECTION_TIME = "Initial Select time"
        self.SECOND_SELECTION_TIME = "Subseq Select time"
        self.WON_ON = " Won On "
        self.DELIVERY_CLICKS = "Delivery clicks"
        self.ITI_CLICKS = "ITI clicks"
        self.DELAY_PAYOFF = "Delay payoff"
        self.DELAY_STREAK = "Delay Consec"
        self.DELAY_PROPORTION = "Delay Proport"
        self.SHORT_PAYOFF = "Short payoff"
        self.SHORT_STREAK = "Short Consec"
        self.SHORT_PROPORTION = "Short Proport"
        self.BANK = "Bank total"
        self.NONOBJECT_CLICKS = "Non-Object Clicks"
        self.FIRST_SELECTION = "First Selected"
        self.FIRST_POSITION = "L or R"
        self.SESSION_TIME = "Session Length"
        self.trialDataList = []
        self.timestamp = time.time()
        self.currentTrial = None
        self.startTime = time.time()
        self.totalTime = 0
        self.columnNames = ["Trial #", "Pre-test clicks", "Delay clicks", "Short clicks", "Non-Object Clicks", "Total clicks", "# Switches", "Trial seconds", "Session Length", "First Selected", 
            "L or R", "Initial Select time", "Subseq Select time", " Won On ", "Delivery clicks", "ITI clicks", "Delay payoff", "Delay Consec", "Delay Proport", "Short payoff", 
            "Short Consec", "Short Proport", "Bank total" ]
    def startSession(self):
        self.startTime = time.time()

    def getTrialNumber(self):
        return len(self.trialDataList)
    def logNonObjectClick(self):
        self.currentTrial.nonObjectClicks = self.currentTrial.nonObjectClicks + 1
    def logClickDuringTrial(self, target):
        self.logObjectClick(target)
        if target != self.lastClickedObject:
            self.currentTrial.switches = self.currentTrial.switches + 1
        self.lastClickedObject = target

    def logPreTestClick(self):
        self.preTestCount = self.preTestCount + 1

    def newTrial(self):
        self.currentTrial = PerTrialData(self.getTrialNumber()+ 1)
        self.trialDataList.append(self.currentTrial)
        self.reStartLocalTimer()
    def logObjectClick(self, object):
        if object == 0:
            self.currentTrial.shortCount = self.currentTrial.shortCount + 1
        elif object == 1:
            self.currentTrial.delayCount = self.currentTrial.delayCount + 1
    def logSubsequentSelection(self):
        if self.currentTrial.secondSelectionTime > 0:
            return
        self.currentTrial.secondSelectionTime = self.elapsedTime()
    def logItiClick(self):
        self.currentTrial.itiClicks = self.currentTrial.itiClicks + 1
    def logPointsDeliveryClick(self):
        self.currentTrial.deliveryClicks = self.currentTrial.deliveryClicks + 1
    def logFirstSelection(self, type, position):
        self.logObjectClick(type)
        self.lastClickedObject = type
        self.currentTrial.initialSelectionTime = self.elapsedTime()
        self.currentTrial.firstSelection = type
        self.currentTrial.firstPosition = position
    
    def reStartLocalTimer(self):
        self.timestamp = time.time()

    def pointsAwarded(self, points, wonOn):
        self.bankTotal = self.bankTotal + points
        self.currentTrial.wonOn = wonOn
        self.currentTrial.payoff = points
        self.currentTrial.total = self.bankTotal
        self.currentTrial.trialLength = self.elapsedTime()
        self.updateStreaks(wonOn)

    def updateStreaks(self, wonOn):
        if wonOn ==0:
            self.updateShortStreak()
        elif wonOn == 1:
            self.updateDelayStreak()

    def updateShortStreak(self):
        self.shortCount = self.shortCount + 1
        self.currentDelayStreak = 0
        self.currentShortStreak = self.currentShortStreak + 1
        if self.currentShortStreak> self.shortConsecutive:
            self.shortConsecutive = self.currentShortStreak

    def updateDelayStreak(self):
        self.delayCount = self.delayCount + 1
        self.currentShortStreak  = 0
        self.currentDelayStreak = self.currentDelayStreak + 1
        if self.currentDelayStreak > self.delayConsecutive:
            self.delayConsecutive = self.currentDelayStreak

    def doneSession(self):
        self.totalTime = time.time() - self.startTime
    def sessionTime(self):
        return self.totalTime
    def elapsedTime(self):
        currentTime = time.time()
        return currentTime - self.timestamp
    def delayStreak(self):
        return self.delayConsecutive
    def shortStreak(self):
        return self.shortConsecutive
    def getDelayWinsSoFar(self):
        return self.delayCount
    def getShortWinsSoFar(self):
        return self.shortCount
    def shortProportion(self):
        return self.getShortWinsSoFar()/self.getTrialNumber()
    def delayProportion(self):
        return self.getDelayWinsSoFar()/self.getTrialNumber()
    def sumDelayClicks(self):
        s = 0
        for trialData in self.trialDataList:
            s= s + trialData.delayCount
        return s
    def sumShortClicks(self):
        s = 0
        for trialData in self.trialDataList:
            s= s + trialData.shortCount
        return s
    def sumNonObjectClicks(self):
        s = 0
        for trialData in self.trialDataList:
            s= s + trialData.nonObjectClicks
        return s
    def sumTotalClicks(self):
        s = 0
        for trialData in self.trialDataList:
            s= s + trialData.getTotalCount()
        return s
    def sumSwitches(self):
        s = 0
        for trialData in self.trialDataList:
            s= s + trialData.switches
        return s
    def sumTrialTime(self):
        s = 0
        for trialData in self.trialDataList:
            s= s + trialData.trialLength
        return s
    def sumInitialSelectionTime(self):
        s = 0
        for trialData in self.trialDataList:
            s= s + trialData.initialSelectionTime
        return s
    def sumSecondSelectionTime(self):
        s = 0
        for trialData in self.trialDataList:
            s= s + trialData.getSecondSelectionTime()
        return s
    def sumDeliveryClicks(self):
        s = 0
        for trialData in self.trialDataList:
            s= s + trialData.deliveryClicks
        return s
    def sumDelayPayoff(self):
        s = 0
        for trialData in self.trialDataList:
            s= s + trialData.delayPayoff()
        return s
    def sumShortPayoff(self):
        s = 0
        for trialData in self.trialDataList:
            s= s + trialData.shortPayoff()
        return s
    def sumItiClicks(self):
        s = 0
        for trialData in self.trialDataList:
            s= s + trialData.itiClicks
        return s
    def getBankTotal(self):
        return self.bankTotal
    def preTestClicks(self):
        return self.preTestCount
    def writeSumRow(self, r):
        self.worksheet.cell(row=r, column=1).value = "TOTALS"
        self.worksheet.cell(row=r, column=2).value = self.preTestClicks()
        self.worksheet.cell(row=r, column=3).value = self.sumDelayClicks()
        self.worksheet.cell(row=r, column=4).value = self.sumShortClicks()
        self.worksheet.cell(row=r, column=5).value = self.sumNonObjectClicks()
        self.worksheet.cell(row=r, column=6).value = self.sumTotalClicks()
        self.worksheet.cell(row=r, column=7).value = self.sumSwitches()
        self.worksheet.cell(row=r, column=8).value = self.sumTrialTime()
        self.worksheet.cell(row=r, column=9).value = self.sessionTime()
        #Fist Selected
        #L or R
        self.worksheet.cell(row=r, column=12).value = self.sumInitialSelectionTime()
        self.worksheet.cell(row=r, column=13).value = self.sumSecondSelectionTime()
        #WonOn
        self.worksheet.cell(row=r, column=15).value = self.sumDeliveryClicks()
        self.worksheet.cell(row=r, column=16).value = self.sumItiClicks()
        self.worksheet.cell(row=r, column=17).value = self.sumDelayPayoff()
        self.worksheet.cell(row=r, column=18).value = self.delayStreak()
        self.worksheet.cell(row=r, column=19).value = self.delayProportion()
        self.worksheet.cell(row=r, column=20).value = self.sumShortPayoff()
        self.worksheet.cell(row=r, column=21).value = self.shortStreak()
        self.worksheet.cell(row=r, column=22).value = self.shortProportion()
        self.worksheet.cell(row=r, column=23).value = self.getBankTotal()
    def writeData(self, reportFilePath, participantData):# export session data to excel
        print("Excel Expert")
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Test Sesstion Data"
        headerfontStyle = styles.Font(color='000000', size=13, bold=True, name='Arial')
        headerfillStyle = styles.PatternFill(start_color='9B9B9B', end_color='9B9B9B', fill_type='solid')
        
        participantData.writeData(self.worksheet)
        i = 1
        for header in self.columnNames:
            self.worksheet.cell(row=5, column=i).value = header
            self.worksheet.cell(row=5, column=i).font = headerfontStyle
            self.worksheet.cell(row=5, column=i).fill =headerfillStyle
            i = i + 1
        
        k = 6
        for trialdata in self.trialDataList:
            self.worksheet.cell(row=k, column=1).value = trialdata.trialNumber
            self.worksheet.cell(row=k, column=2).value = self.preTestClicks()
            self.worksheet.cell(row=k, column=3).value = trialdata.delayCount
            self.worksheet.cell(row=k, column=4).value = trialdata.shortCount
            self.worksheet.cell(row=k, column=5).value = trialdata.nonObjectClicks
            self.worksheet.cell(row=k, column=6).value = trialdata.getTotalCount()
            self.worksheet.cell(row=k, column=7).value = trialdata.switches
            self.worksheet.cell(row=k, column=8).value = trialdata.trialLength
            self.worksheet.cell(row=k, column=9).value = self.sessionTime()
            firstSelection = "delay"
            if trialdata.firstSelection == 0:
                firstSelection = "short"
            self.worksheet.cell(row=k, column=10).value = firstSelection
            firstPosition = "R"
            if trialdata.firstPosition == 0:
                firstPosition = "L"
            self.worksheet.cell(row=k, column=11).value = firstPosition
            self.worksheet.cell(row=k, column=12).value = trialdata.initialSelectionTime
            self.worksheet.cell(row=k, column=13).value = trialdata.getSecondSelectionTime()
            wonOnText = "delay"
            if trialdata.wonOn == 0:
                wonOnText = "short"
            self.worksheet.cell(row=k, column=14).value = wonOnText
            self.worksheet.cell(row=k, column=15).value = trialdata.deliveryClicks
            self.worksheet.cell(row=k, column=16).value = trialdata.itiClicks
            self.worksheet.cell(row=k, column=17).value = trialdata.delayPayoff()
            self.worksheet.cell(row=k, column=18).value = self.delayStreak()
            self.worksheet.cell(row=k, column=19).value = self.delayProportion()
            self.worksheet.cell(row=k, column=20).value = trialdata.shortPayoff()
            self.worksheet.cell(row=k, column=21).value = self.shortStreak()
            self.worksheet.cell(row=k, column=22).value = self.shortProportion()
            self.worksheet.cell(row=k, column=23).value = trialdata.total
            k = k + 1
        self.writeSumRow(k)
        self.workbook.save(reportFilePath + "/"+ participantData.txt_ID.text() + ".xlsx")

class ParticipantWindow(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.mainwin = parent
        self.setWindowTitle("New Participant")
        self.setFont(QFont("Arial", 12, -1, False))
        lbl_id = QLabel("ID:")
        lbl_age = QLabel("Age:")
        lbl_gender = QLabel("Gender")
        lbl_handedness = QLabel("Handedness")
        lbl_education = QLabel("Education")
        lbl_race = QLabel("Race/Ethnicity")
        lbl_learning = QLabel("Learning or Mental Disorder")

        self.txt_ID = QLineEdit()
        random_string = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
        self.txt_ID.setText(random_string)
        self.txt_Age = QSpinBox()
        self.combo_Gender = QComboBox()
        self.combo_Gender.addItem("Male")
        self.combo_Gender.addItem("Female")
        self.combo_Handedness = QComboBox()
        self.combo_Handedness.addItem("Left")
        self.combo_Handedness.addItem("Right")
        self.combo_Education = QComboBox()

        for i in range(1, 21) :
            item_str = str(i)
            if i == 12:
                item_str = item_str + "(HSD or GED)"
            elif i == 14:
                item_str = item_str + "(AA)"
            elif i == 16:
                item_str = item_str + "(BA)"
            elif i == 18:
                item_str = item_str + "(MA)"
            elif i == 20:
                item_str = item_str + "(Doctorate)"
            self.combo_Education.addItem(item_str)

        self.combo_Race = QComboBox()
        self.combo_Race.addItem("Decline to State")
        self.combo_Race.addItem("Asian/Asian American")
        self.combo_Race.addItem("African/African American")
        self.combo_Race.addItem("Alaskan Native")
        self.combo_Race.addItem("Caucasian/Caucasian American")
        self.combo_Race.addItem("Latino/Latino American")
        self.combo_Race.addItem("Native American")
        self.combo_Race.addItem("Multiracial")

        self.combo_Learning = QComboBox()
        self.combo_Learning.addItem("Decline to State")
        self.combo_Learning.addItem("Anxiety Disorder")
        self.combo_Learning.addItem("Learning Disorder")
        self.combo_Learning.addItem("Mood Disorder")
        self.combo_Learning.addItem("Neurological Disorder/Brain Injury")
        self.combo_Learning.addItem("Personality Disorder")
        self.combo_Learning.addItem("Psychotic Disorder")
        self.combo_Learning.addItem("More than one disorder present")

        form_lay = QFormLayout()
        form_lay.addRow(lbl_id,self.txt_ID)
        form_lay.addRow(lbl_age,self.txt_Age)
        form_lay.addRow(lbl_gender,self.combo_Gender)
        form_lay.addRow(lbl_handedness,self.combo_Handedness)
        form_lay.addRow(lbl_education,self.combo_Education)
        form_lay.addRow(lbl_race,self.combo_Race)
        form_lay.addRow(lbl_learning,self.combo_Learning)
        
        dialog_lay = QVBoxLayout()
        btn_lay = QHBoxLayout()
        
        btn_ok = QPushButton("Ok")
        btn_ok.clicked.connect(self.setNewUserInfo)
        btn_cancel = QPushButton("Cancel")
        btn_lay.addStretch()
        btn_lay.addWidget(btn_ok)
        btn_lay.addWidget(btn_cancel)
        btn_cancel.clicked.connect(self.close)
        dialog_lay.addLayout(form_lay)
        dialog_lay.addLayout(btn_lay)
        self.setLayout(dialog_lay)
        self.setFixedSize(600, 300)

    def writeData(self, sheet):
        headerStyle= styles.PatternFill(start_color='9B9B9B', end_color='9B9B9B', fill_type='solid')
        headerFont = styles.Font(color='000000', size=13, bold=True, name='Arial')
        sheet.cell(row=1, column=1).value = "A-GRAF"
        sheet.cell(row=1, column=1).font = headerFont
        sheet.cell(row=1, column=1).fill = headerStyle
        sheet.cell(row=1, column=3).value = "ID"
        sheet.cell(row=1, column=3).font = headerFont
        sheet.cell(row=1, column=3).fill = headerStyle
        sheet.cell(row=2, column=3).value = self.txt_ID.text()
        sheet.cell(row=1, column=4).value = "Age"
        sheet.cell(row=1, column=4).font = headerFont
        sheet.cell(row=1, column=4).fill = headerStyle
        sheet.cell(row=2, column=4).value = self.txt_Age.value()
        sheet.cell(row=1, column=5).value = "Gender"
        sheet.cell(row=1, column=5).font = headerFont
        sheet.cell(row=1, column=5).fill = headerStyle
        sheet.cell(row=2, column=5).value = self.combo_Gender.currentText()
        sheet.cell(row=1, column=6).value = "Handedness"
        sheet.cell(row=1, column=6).font = headerFont
        sheet.cell(row=1, column=6).fill = headerStyle
        sheet.cell(row=2, column=6).value = self.combo_Handedness.currentText()
        sheet.cell(row=1, column=7).value = "Education"
        sheet.cell(row=1, column=7).font = headerFont
        sheet.cell(row=1, column=7).fill = headerStyle
        sheet.cell(row=2, column=7).value = self.combo_Education.currentText()
        sheet.cell(row=1, column=8).value = "Race"
        sheet.cell(row=1, column=8).font = headerFont
        sheet.cell(row=1, column=8).fill = headerStyle
        sheet.cell(row=2, column=8).value = self.combo_Race.currentText()
        sheet.cell(row=1, column=9).value = "Disorder"
        sheet.cell(row=1, column=9).font = headerFont
        sheet.cell(row=1, column=9).fill = headerStyle
        sheet.cell(row=2, column=9).value = self.combo_Learning.currentText()
    def setNewUserInfo(self):
        self.mainwin.userId = self.txt_ID.text()
        self.mainwin.userAge = self.txt_Age.value()
        self.mainwin.userGender = self.combo_Gender.currentText()
        self.mainwin.userHandedness = self.combo_Handedness.currentText()
        self.mainwin.userEducation = self.combo_Education.currentText()
        self.mainwin.userRace = self.combo_Race.currentText()
        self.mainwin.userLearning = self.combo_Learning.currentText()
        self.mainwin.btn_user.setText(self.txt_ID.text())
        self.close()

class SelectableObject(QLabel):
    my_press_signal = pyqtSignal(QMouseEvent)
    def __init__(self, oType, position, mainwin) -> None:
        super().__init__(parent=mainwin)
        self.objectType = oType
        self.position = position
        self.isSelectedObject = False
        self.setObjectType(self.objectType)
        self.mousePressEvent = lambda event: self.mouse_pressed(event)
    def mouse_pressed(self, event):
        self.my_press_signal.emit(event)
    def setObjectType(self, oType):
        self.objectType = oType
        self.triangle = QPixmap(":images/triangle-lines.gif")
        self.star = QPixmap(":images/star-lines.gif")
        self.triangle_selected = QPixmap(":images/triangle-lines-selected.gif")
        self.star_selected = QPixmap(":images/star-lines-selected.gif")
        self.triangle_selected_point = QPixmap(":images/triangle-lines-selected-point-delivery.gif")
        self.star_selected_point = QPixmap(":images/star-lines-selected-point-delivery.gif")
        if self.objectType == 0:
            self.setVisible(True)
            self.setPixmap(self.triangle)
        elif self.objectType == 1:
            self.setVisible(True)
            self.setPixmap(self.star)
        elif self.objectType == 2:
            self.setVisible(False)
        # self.isSelectedObject = False

    def setPointDeliveryState(self):
        if self.objectType == 0:
            self.setPixmap(self.triangle_selected_point)
        elif self.objectType == 1:
            self.setPixmap(self.star_selected_point)
        
    def getPosition(self):
        return self.position
    
    def setOtherObject(self, other):
        self.other = other

    def select(self):
        if self.other.isSelectedObject:
            self.other.deSelect()
        if self.objectType == 0:
            self.setVisible(True)
            self.setPixmap(self.triangle_selected)
            self.update()
        elif self.objectType == 1:
            self.setVisible(True)
            self.setPixmap(self.star_selected)
            self.update()
        self.isSelectedObject = True
    
    def deSelect(self):
        if self.objectType == 0:
            self.setVisible(True)
            self.setPixmap(self.triangle)
        elif self.objectType == 1:
            self.setVisible(True)
            self.setPixmap(self.star)
        self.isSelectedObject = False
        
    def getDelay(self, log, config):
        if self.objectType == 0:
            return config.getShortDelay(log.shortCount)
        elif self.objectType == 1:
            return config.getDelayedDelay(log.delayCount)
        else:
            return 0
        
    def getPoints(self, log, config):
        if self.objectType ==0:
            return config.getShortPoints(log.shortCount)
        elif self.objectType == 1:
            return config.getDelayedPoints(log.delayCount)
        else:
            return 0
    def getITI(self, config):
        if self.objectType == 0:
            return config.getShortITI()
        elif self.objectType == 1:
            return config.getDelayedITI()
    def canBeSelected(self,allowSwitchBack, noSwitches):
        if self.isSelectedObject:
            return False
        elif noSwitches and self.other.isSelectedObject:
            return False
        elif self.objectType == 2:
            return False
        elif self.objectType == 1 and self.other.isSelectedObject and not allowSwitchBack:
            return False
        else:
            return True
    def tryToSelect(self,allowSwitchBack, noSwitches):
        if not self.canBeSelected(allowSwitchBack, noSwitches):
            return 
        else:
            self.select()
            return True
class AlartLight(QLabel):
    def __init__(self):
        super().__init__()
        self.pixmap_red = QPixmap(':red.png')
        self.pixmap_green = QPixmap(':green.png')
        self.setFixedHeight(120)
    def setInitialState(self):
        self.setPixmap(self.pixmap_red)

    def switchStatus(self, isEnabled):
        if isEnabled:
            self.setPixmap(self.pixmap_green)
        else:
            self.setPixmap(self.pixmap_red)

class SessionWindow(QDialog):
    def __init__(self, parent, type):
        super().__init__(parent)
        self.setStyleSheet("background-color: white;")
        self.config = parent.window_configuration
        self.participant = parent.window_new_participant
        self.sessionType = type
        self.sessionAlertTimer = QTimer()
        self.sessionAlertTimer.setSingleShot(True)
        self.sessionStatus = 0
        self.statusLight = AlartLight()
        self.lay_session = QVBoxLayout()
        self.lbl_title = QLabel()
        self.lbl_title.setText("This is a Practice Session")
        self.lbl_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_title.setFixedHeight(60)
        self.lbl_title.setFont(QFont("Arial", 20, -1, False))
        self.lbl_title.setStyleSheet("background-color: blue; color:white;")
        self.lay_session.addWidget(self.lbl_title)
        self.setLayout(self.lay_session)
        self.setSessionWidget()
        self.setCountingWidget()
        self.setCongratsWidget()
        self.lay_session.addWidget(self.sessionWidget)
        self.runSession(self.sessionType)
    
    def setCongratulationText(self, finalPoints):
        self.lbl_congrats.setText("Congratulations! You won " + str(finalPoints)+ " points!")

    def setCongratsWidget(self):
        self.congratsWidget = QWidget()
        self.setStyleSheet("background-color: white;")
        lay_main = QVBoxLayout()
        self.lbl_congrats = QLabel()
        self.lbl_congrats.setFont(QFont("Arial", 22, -1, False))
        self.lbl_congrats.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay_bottom = QHBoxLayout()
        lay_bottom.addStretch()
        btn_close = QPushButton("Close")
        lay_bottom.addWidget(btn_close)
        btn_close.clicked.connect(self.close)
        lay_bottom.addStretch()
        lay_main.addStretch()
        lay_main.addWidget(self.lbl_congrats)
        lay_main.addStretch()
        lay_main.addLayout(lay_bottom)
        self.congratsWidget.setLayout(lay_main)

    def setCountingWidget(self):
        self.countingWidget =QWidget()
        lay_main = QVBoxLayout()
        self.countingWidget.setLayout(lay_main)
        lay_bottom = QHBoxLayout()
        lay_bottom.addStretch()
        btn_close = QPushButton("Close")
        lay_bottom.addWidget(btn_close)
        btn_close.clicked.connect(self.close)
        lay_bottom.addStretch()
        self.lbl_counting = QLabel()
        self.lbl_counting.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_counting.setFont(QFont("Arial", 100, 50, False))
        lay_main.addWidget(self.lbl_counting)
        lay_main.addLayout(lay_bottom)

    def setSessionWidget(self):
        self.sessionWidget = QWidget()
        lay_main = QVBoxLayout()
        self.sessionWidget.setLayout(lay_main)
        lay_status = QHBoxLayout()
        lay_status.addStretch(1)
        lay_status.addWidget(self.statusLight)
        lay_status.addStretch(1)
        
        lay_main.addLayout(lay_status)
        group_object = QWidget()
        self.object_left = SelectableObject(0, 0, self)
        self.object_right = SelectableObject(1, 1, self)
        self.object_left.setOtherObject(self.object_right)
        self.object_left.my_press_signal.connect(self.objectPressed)
        self.object_right.setOtherObject(self.object_left)
        self.object_right.my_press_signal.connect(self.objectPressed)
        lay_group = QHBoxLayout()
        lay_left_v = QVBoxLayout()
        lay_right_v = QVBoxLayout()
        lay_left = QHBoxLayout()
        lay_right = QHBoxLayout()
        lay_left.addStretch()
        lay_left.addWidget(self.object_left)
        lay_left.addStretch()
        lay_right = QHBoxLayout()
        lay_right.addStretch()
        lay_right.addWidget(self.object_right)
        lay_right.addStretch()
        lay_left_v.addStretch()
        lay_left_v.addLayout(lay_left)
        lay_left_v.addStretch()
        lay_right_v.addStretch()
        lay_right_v.addLayout(lay_right)
        lay_right_v.addStretch()
        lay_group.addLayout(lay_left_v)
        lay_group.addLayout(lay_right_v)
        group_object.setLayout(lay_group)
        lay_main.addWidget(group_object)
        lay_bottom = QHBoxLayout()
        self.lbl_point = SessionPoints()
        self.lbl_point.setPoints(0)
        lay_bottom.addStretch(1)
        lay_bottom.addWidget(self.lbl_point)
        lay_bottom.addStretch(1)
        lay_main.addLayout(lay_bottom)
        
    def runSession(self, type):
        self.sessionType = type
        if type == 0:
            self.lbl_title.setText("This is a Practice Session")
            self.log = DataLog(self.config.getNumberOfPracticeTrials() * 2)
        elif type ==1:
            self.lbl_title.setText("")
            self.log = DataLog(self.config.getNumberOfPracticeTrials())
        self.lbl_title.update()
        self.setUpTrial()
        self.PreTestState()
        
    def setUpTrial(self):
        if self.sessionType == 0:# practice session type
            if self.log.getTrialNumber() < self.config.getNumberOfPracticeTrials():
                self.randomizePlacement(self.config.getFirstPracticeObject(), 2)
            else:
                self.randomizePlacement(self.config.getSecondPracticeObject(), 2)
        elif self.sessionType == 1:# test session type
            self.randomizePlacement(0, 1)

    def PreTestState(self):
        self.sessionStatus = 0
        self.statusLight.setInitialState()
        self.log.newTrial()
        self.lbl_point.resetPoints()
        self.sessionAlertTimer.singleShot(1000 * self.config.getPracticeDelay(), self.TrialBeginState)

    def TrialBeginState(self):
        self.sessionStatus = 1
        self.log.startSession()
        self.statusLight.switchStatus(True)
    
    def TrialSelectionState(self):
        self.sessionStatus = 2
        self.selectedObject.select()
        self.log.logFirstSelection(self.selectedObject.objectType, self.selectedObject.position)
        self.sessionAlertTimer.singleShot(1000 * self.selectedObject.getDelay(self.log,self.config), self.PointsAwardState)

    def PointsAwardState(self):
        self.sessionStatus = 3
        self.statusLight.switchStatus(False)
        self.selectedObject.setPointDeliveryState()
        self.payoff = self.selectedObject.getPoints(self.log,self.config)
        self.log.pointsAwarded(self.payoff,self.selectedObject.objectType)
        self.lbl_point.countUp(self.payoff, self.config.getPointsDelay())
        self.lbl_point.startBlinking()
        
        self.sessionAlertTimer.singleShot(1000 * self.config.getPointsDelay(), self.ItiState)
    def ItiState(self):
        self.sessionStatus = 4
        self.sessionAlertTimer.singleShot(1000 * self.selectedObject.getITI(self.config), self.EndState)

    def isFinished(self):
        if self.sessionType == 0:
            return self.log.getTrialNumber() == self.config.getNumberOfPracticeTrials() * 2
        elif self.sessionType == 1:
            return self.log.getTrialNumber() == self.config.getNumberOfTestTrials()
    def startCounting(self):
        while self.countnum >= 0:
            self.lbl_counting.setText(str(self.countnum))
            time.sleep(1)
            self.countnum = self.countnum - 1
    def ready4Test(self):
        self.lbl_title.setText("The practice session is now over. GET READY TO START THE GAME!")
        self.sessionWidget.setVisible(False)
        self.lay_session.addWidget(self.countingWidget)
        self.countingWidget.setVisible(True)
        self.lbl_point.stopBlinking()
        self.countnum = 10
        self.count4TestThread = threading.Thread(target=self.startCounting)
        self.count4TestThread.start()
        self.sessionAlertTimer.singleShot(1000 * 10, self.startTestSession)

    def startTestSession(self):
        self.countingWidget.setVisible(False)
        self.sessionWidget.setVisible(True)
        self.runSession(1)
        
    def EndState(self):
        if self.isFinished():# session finished
            if self.sessionType == 0:
                self.ready4Test()
            elif self.sessionType == 1:
                self.log.doneSession()
                self.sessionWidget.setVisible(False)
                self.lay_session.addWidget(self.congratsWidget)
                self.congratsWidget.setVisible(True)
                self.setCongratulationText(self.lbl_point.getPoints())
                background_thread = threading.Thread(target=self.log.writeData, args=(self.config.txt_directory.text(), self.participant))
                background_thread.start()
        else:
            self.sessionStatus = 5
            self.lbl_point.stopBlinking()
            self.setUpTrial()
            self.log.newTrial()
            self.TrialBeginState()
    
    def resetPoint(self):
        self.lbl_point.setPoints(0)

    def randomizePlacement(self, shapeType, otherShapeType):
        random_value = random.randint(0, 1)
        if random_value == 0:
            self.object_left.setObjectType(shapeType)
            self.object_right.setObjectType(otherShapeType)
        else:
            self.object_right.setObjectType(shapeType)
            self.object_left.setObjectType(otherShapeType)

    def keyPressEvent(self, event: QKeyEvent):
        if event.key() == 16777216:
                self.close()
        if self.sessionStatus == 0:
            self.log.logPreTestClick()
        elif self.sessionStatus == 1:
            if event.key() == 48:
                if self.object_right.objectType < 2:
                    self.selectedObject = self.object_right
                    self.TrialSelectionState()
            elif event.key() == 49:
                if self.object_left.objectType < 2:
                    self.selectedObject = self.object_left
                    self.TrialSelectionState()
            else:
                self.log.logNonObjectClick()
        elif self.sessionStatus == 2:
            if event.key() == 48:
                if self.object_right.objectType< 2 and not self.object_right.isSelectedObject:
                    if self.object_right.tryToSelect(self.config.isAllowSwitchBack(), False):
                        self.selectedObject = self.object_right
                        self.log.logSubsequentSelection()
            elif event.key() == 49:
                if self.object_left.objectType< 2 and not self.object_left.isSelectedObject:
                    if self.object_left.tryToSelect(self.config.isAllowSwitchBack(), False):
                        self.selectedObject = self.object_left
                        self.log.logSubsequentSelection()
            else:
                self.log.logNonObjectClick()
        elif self.sessionStatus == 3:
            self.log.logPointsDeliveryClick()
        elif self.sessionStatus == 4:
            self.log.logItiClick()
        
    # Override the mouse press event method
    def objectPressed(self,event):
        print("Mouse Pressed")
        if self.sender():
            if self.sessionStatus == 0:
                self.log.logPreTestClick()
            elif self.sessionStatus == 1:
                if self.sender().position == 0:
                    if self.sender().objectType < 2:
                        self.selectedObject = self.sender()
                        self.TrialSelectionState()
                elif self.sender().position == 1:
                    if self.sender().objectType < 2:
                        self.selectedObject = self.sender()
                        self.TrialSelectionState()
            elif self.sessionStatus == 2:
                if self.sender().position == 0 and not self.sender().isSelectedObject:
                    if self.sender().objectType< 2:
                        if self.sender().tryToSelect(self.config.isAllowSwitchBack(), False):
                            self.selectedObject = self.sender()
                            self.log.logSubsequentSelection()
                elif self.sender().position == 1 and not self.sender().isSelectedObject:
                    if self.sender().objectType< 2:
                        if self.sender().tryToSelect(self.config.isAllowSwitchBack(), False):
                            self.selectedObject = self.sender()
                            self.log.logSubsequentSelection()
                else:
                    self.log.logNonObjectClick()
            elif self.sessionStatus == 3:
                self.log.logPointsDeliveryClick()
            elif self.sessionStatus == 4:
                self.log.logItiClick()
class SessionPoints(QLabel):
    def __init__(self):
        super().__init__()
        self.isNormal = True
        self.setFont(QFont("Arial", 30, -1, False))
        self.normalStyle = "border: 5px solid;border-radius: 10px; background-color: white; color:black;"
        self.blinkStyle = "border: 5px solid;border-radius: 10px; background-color: rgb(6, 167, 255); color:blue;"
        self.setStyleSheet(self.normalStyle)
        self.setFixedSize(QSize(80,80))
        self.setAlignment(Qt.AlignmentFlag.AlignCenter)

    def setPoints(self, points):
        self.setText(str(points))

    def setBlinking(self):
        while self.isBlinking:
            if self.isNormal :
                self.setStyleSheet(self.blinkStyle)
                self.isNormal = False
            else :
                self.setStyleSheet(self.normalStyle)
                self.isNormal = True
            time.sleep(0.45)
            # self.update()
    def setNormal(self):
        self.setStyleSheet(self.normalStyle)
        self.isNormal = True

    def countPoint(self):
        while self.isCounting:
            points = self.getPoints()
            if points >= self.target:
                self.isCounting = False
            else:
                self.setPoints(points + 1)
            time.sleep(self.timePerCountMillis)
            self.update()
    def countUp(self, payoff, pointsDelay):
        if payoff< 1:
            return
        self.target = self.getPoints() + payoff
        # self.deadline = time.time() + pointsDelay
        if payoff > 1:
            self.timePerCountMillis = pointsDelay/(payoff - 1)
        else:
            self.timePerCountMillis = pointsDelay
        self.isCounting = True
        self.countThread = threading.Thread(target=self.countPoint)
        self.countThread.start()

    def stopCounting(self):
        self.isCounting = False

    def resetPoints(self):
        self.setText("0")

    def startBlinking(self):
        self.isBlinking = True
        self.blinkThread = threading.Thread(target=self.setBlinking)
        self.blinkThread.start()

    def getPoints(self):
        return int(self.text())
    
    def stopBlinking(self):
        self.isBlinking = False
        self.setNormal()
class ConfigWindow(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.mainwin = parent
        self.setFont(QFont("Arial", 13, -1, False))
        self.setWindowTitle("A-GRAF Settings")
        main_lay = QVBoxLayout()
        testsession_lay = QVBoxLayout()
        trial_num_lay = QHBoxLayout()
        directory_lay = QHBoxLayout()
        switch_lay = QHBoxLayout()
        lbl_tnum = QLabel("Number of trials:")
        self.spin_testtrial_num = QSpinBox()
        self.spin_testtrial_num.setValue(20)
        lbl_directory = QLabel("Store report files in:")
        self.txt_directory = QLineEdit()
        app_directory = QCoreApplication.applicationDirPath()
        self.txt_directory.setText(app_directory)
        self.chk_allowBack = QCheckBox()
        self.chk_allowBack.setText("Allow Switch Back")
        trial_num_lay.addWidget(lbl_tnum)
        trial_num_lay.addWidget(self.spin_testtrial_num)
        directory_lay.addWidget(lbl_directory)
        directory_lay.addWidget(self.txt_directory)
        btn_browse = QPushButton("Browse...")
        btn_browse.clicked.connect(self.open_directory_dialog)
        directory_lay.addWidget(btn_browse)
        
        switch_lay.addWidget(self.chk_allowBack)

        testsession_lay.addLayout(trial_num_lay)
        testsession_lay.addLayout(directory_lay)
        testsession_lay.addLayout(switch_lay)
        group_test_config = QGroupBox("Test Session Settings")
        group_test_config.setLayout(testsession_lay)

        practice_lay = QVBoxLayout()
        group_practice_config = QGroupBox("Practice Session Settings")
        group_practice_config.setLayout(practice_lay)
        self.radio_triagle = QRadioButton("Display Trianlges (short) first")
        self.radio_triagle.setChecked(True)
        self.radio_star = QRadioButton("Display Stars (delayed) first")
        lbl_trial_num = QLabel("Number of trials per object:")
        self.spin_practicetrial_num = QSpinBox()
        self.spin_practicetrial_num.setValue(5)
        trial_lay = QHBoxLayout()
        
        trial_lay.addWidget(lbl_trial_num)
        trial_lay.addWidget(self.spin_practicetrial_num)
        practice_lay.addWidget(self.radio_triagle)
        practice_lay.addWidget(self.radio_star)
        practice_lay.addLayout(trial_lay)

        polygon_lay = QHBoxLayout()
        triangel_group = QGroupBox("Triangle (short) Payoff")
        triangle_lay = QVBoxLayout()
        triangel_group.setLayout(triangle_lay)
        triangle_formlay = QFormLayout()
        star_formlay = QFormLayout()

        lbl_initialPayoff = QLabel("Initial Payoff:")
        self.spin_initialPayoff = QSpinBox()
        self.spin_initialPayoff.valueChanged.connect(self.setTriangleTableContent)
        triangle_formlay.addRow(lbl_initialPayoff, self.spin_initialPayoff)
        lbl_incrementalPayoff = QLabel("Incremental Payoff:")
        self.spin_incrementalPayoff = QSpinBox()
        self.spin_incrementalPayoff.valueChanged.connect(self.setTriangleTableContent)
        triangle_formlay.addRow(lbl_incrementalPayoff, self.spin_incrementalPayoff)
        lbl_delay = QLabel("Delay (seconds):")
        self.spin_delay_triangle = QSpinBox()
        self.spin_delay_triangle.valueChanged.connect(self.setTriangleTableContent)
        triangle_formlay.addRow(lbl_delay, self.spin_delay_triangle)
        lbl_delayincre = QLabel("Delay increment:")
        self.spin_delayincre_triangle = QSpinBox()
        self.spin_delayincre_triangle.valueChanged.connect(self.setTriangleTableContent)
        lbl_interval = QLabel("Inter-Trial Interval:")
        self.spin_interval = QSpinBox()
        triangle_formlay.addRow(lbl_delayincre, self.spin_delayincre_triangle)
        triangle_formlay.addRow(lbl_interval, self.spin_interval)
        self.chk_geoPayoffIncre = QCheckBox()
        self.chk_geoPayoffIncre.setText("Geometric Payoff Increase")
        self.chk_geoPayoffIncre.stateChanged.connect(self.setTriangleTableContent)
        radio_payoffpercent = QRadioButton("100%")
        radio_payoffpercent.setChecked(True)
        radio_payoffpercent.setEnabled(False)
        self.chk_geoDelayIncre = QCheckBox("Geometric Delay Increase")
        self.chk_geoDelayIncre.stateChanged.connect(self.setTriangleTableContent)
        radio_delaypercent = QRadioButton("100%")
        radio_delaypercent.setChecked(True)
        radio_delaypercent.setEnabled(False)
        triangle_formlay.addRow(self.chk_geoPayoffIncre, radio_payoffpercent)
        triangle_formlay.addRow(self.chk_geoDelayIncre, radio_delaypercent)
        self.table_Triangle = QTableWidget()
        self.table_Triangle.setColumnCount(6)
        self.table_Triangle.setRowCount(2)
        self.table_Triangle.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)

        triangle_sitem = QTableWidgetItem("Seconds")
        triangle_pitem = QTableWidgetItem("Points") 
        self.table_Triangle.horizontalHeader().hide()
        self.table_Triangle.setVerticalHeaderItem(0,triangle_sitem)
        self.table_Triangle.setVerticalHeaderItem(1,triangle_pitem)
        triangle_lay.addLayout(triangle_formlay)
        triangle_lay.addWidget(self.table_Triangle)
        self.table_Triangle.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

        star_group = QGroupBox("Star (delayed) Payoff")
        star_lay = QVBoxLayout()
        star_group.setLayout(star_lay)
        lbl_initialPayoff_star = QLabel("Initial Payoff:")
        self.spin_initialPayoff_star = QSpinBox()
        self.spin_initialPayoff_star.valueChanged.connect(self.setStarTableContent)
        star_formlay.addRow(lbl_initialPayoff_star, self.spin_initialPayoff_star)
        lbl_incrementalPayoff_star = QLabel("Incremental Payoff:")
        self.spin_incrementalPayoff_star = QSpinBox()
        self.spin_incrementalPayoff_star.valueChanged.connect(self.setStarTableContent)
        star_formlay.addRow(lbl_incrementalPayoff_star, self.spin_incrementalPayoff_star)
        lbl_delay_star = QLabel("Delay (seconds):")
        self.spin_delay_star = QSpinBox()
        self.spin_delay_star.valueChanged.connect(self.setStarTableContent)
        star_formlay.addRow(lbl_delay_star, self.spin_delay_star)
        lbl_delayincre_star = QLabel("Delay increment:")
        self.spin_delayincre_star = QSpinBox()
        self.spin_delayincre_star.valueChanged.connect(self.setStarTableContent)
        lbl_interval_star = QLabel("Inter-Trial Interval:")
        self.spin_interval_star = QSpinBox()
        self.spin_interval_star.valueChanged.connect(self.setStarTableContent)
        star_formlay.addRow(lbl_delayincre_star, self.spin_delayincre_star)
        star_formlay.addRow(lbl_interval_star, self.spin_interval_star)
        self.chk_geoPayoffIncre_star = QCheckBox()
        self.chk_geoPayoffIncre_star.setText("Geometric Payoff Increase")
        self.chk_geoPayoffIncre_star.stateChanged.connect(self.setStarTableContent)
        radio_payoffpercent_star = QRadioButton("100%")
        radio_payoffpercent_star.setChecked(True)
        radio_payoffpercent_star.setEnabled(False)
        self.chk_geoDelayIncre_star = QCheckBox("Geometric Delay Increase")
        self.chk_geoDelayIncre_star.stateChanged.connect(self.setStarTableContent)
        radio_delaypercent_star = QRadioButton("100%")
        radio_delaypercent_star.setChecked(True)
        radio_delaypercent_star.setEnabled(False)
        star_formlay.addRow(self.chk_geoPayoffIncre_star, radio_payoffpercent_star)
        star_formlay.addRow(self.chk_geoDelayIncre_star, radio_delaypercent_star)
        self.table_star = QTableWidget()
        self.table_star.setColumnCount(6)
        self.table_star.setRowCount(2)
        self.table_star.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        
        star_sitem = QTableWidgetItem("Seconds")
        star_pitem = QTableWidgetItem("Points")
        self.table_star.horizontalHeader().hide()
        self.table_star.setVerticalHeaderItem(0, star_sitem)
        self.table_star.setVerticalHeaderItem(1, star_pitem)

        star_lay.addLayout(star_formlay)
        star_lay.addWidget(self.table_star)
        self.table_star.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

        polygon_lay.addWidget(triangel_group)
        polygon_lay.addWidget(star_group)

        foot_lay = QHBoxLayout()
        time_lay = QFormLayout()
        lbl_pretestdelay = QLabel("Pre-test delay")
        self.spin_pretestdelay = QSpinBox()
        lbl_pointdelivery = QLabel("Point delivery time:")
        self.spin_pointdelivery = QSpinBox()
        time_lay.addRow(lbl_pretestdelay, self.spin_pretestdelay)
        time_lay.addRow(lbl_pointdelivery, self.spin_pointdelivery)
        foot_lay.addLayout(time_lay)
        foot_lay.addStretch()
        btn_ok = QPushButton("Ok")
        btn_ok.clicked.connect(self.setConfig)
        foot_lay.addWidget(btn_ok)
        main_lay.addWidget(group_test_config)
        main_lay.addWidget(group_practice_config)
        main_lay.addLayout(polygon_lay)
        main_lay.addLayout(foot_lay)
        self.spin_pretestdelay.setValue(3)
        self.spin_pointdelivery.setValue(3)
        self.spin_initialPayoff.setValue(3)
        self.spin_incrementalPayoff.setValue(3)
        self.spin_delay_triangle.setValue(2)
        self.spin_delayincre_triangle.setValue(2)
        self.spin_interval.setValue(4)

        self.spin_initialPayoff_star.setValue(5)
        self.spin_incrementalPayoff_star.setValue(5)
        self.spin_delay_star.setValue(5)
        self.spin_delayincre_star.setValue(5)
        self.spin_interval_star.setValue(3)
        self.setLayout(main_lay)

    def getDelayedITI(self):
        return self.spin_interval_star.value()
    
    def getShortITI(self):
        return self.spin_interval.value()
    
    def getPointsDelay(self):
        return self.spin_pointdelivery.value()
    
    def getPracticeDelay(self):
        return self.spin_delay_triangle.value()
    
    def getNumberOfTestTrials(self):
        return self.spin_testtrial_num.value()
    def getNumberOfPracticeTrials(self):
        return self.spin_practicetrial_num.value()
    
    def getFirstPracticeObject(self):
        if self.radio_triagle.isChecked():
            return 0
        else:
            return 1
    def isAllowSwitchBack(self):
        return self.chk_allowBack.isChecked()
    def getSecondPracticeObject(self):
        if self.radio_triagle.isChecked():
            return 1
        else:
            return 0
    def calculateCurrentValue(self, count, initial, increment, geometric, callFrom):
        perc = 0
        offPerc = False
        if geometric:
            if callFrom ==1:
                offPerc = self.chk_geoDelayIncre.isChecked()
                if offPerc:
                    perc = 50
                else:
                    perc = 100
            elif callFrom == 2:
                offPerc = self.chk_geoPayoffIncre.isChecked()
                if offPerc:
                    perc = 50
                else:
                    perc = 100
            elif callFrom == 3:
                offPerc = self.chk_geoDelayIncre_star.isChecked()
                if offPerc:
                    perc = 50
                else:
                    perc = 100
            elif callFrom == 4:
                offPerc = self.chk_geoPayoffIncre_star.isChecked()
                if offPerc:
                    perc = 50
                else:
                    perc = 100
            count = count + 1
            sum = initial
            for i in range(2, count):
                if perc == 100:
                    sum = sum * 2
            return sum
        
        return initial + count * increment
    
    def getShortDelay(self, timesAlreadyRewarded):
        initial = self.spin_delay_triangle.value()
        increment = self.spin_delayincre_triangle.value()
        geometric = self.chk_geoPayoffIncre.isChecked()
        return self.calculateCurrentValue(timesAlreadyRewarded, initial, increment, geometric,1)
    
    def getShortPoints(self, timesAlreadyRewarded):
        initial = self.spin_initialPayoff.value()
        increment = self.spin_incrementalPayoff.value()
        geometric = self.chk_geoDelayIncre.isChecked()
        return self.calculateCurrentValue(timesAlreadyRewarded,initial,increment, geometric, 2)
    
    def getDelayedDelay(self, timesAlreadyRewarded):
        initial = self.spin_delay_star.value()
        increment = self.spin_delayincre_star.value()
        geometric = self.chk_geoDelayIncre_star.isChecked()
        return self.calculateCurrentValue(timesAlreadyRewarded, initial, increment, geometric, 3)
    
    def getDelayedPoints(self, timesAlreadyRewarded):
        initial = self.spin_initialPayoff_star.value()
        increment = self.spin_incrementalPayoff_star.value()
        geometric = self.chk_geoPayoffIncre_star.isChecked()
        return self.calculateCurrentValue(timesAlreadyRewarded, initial, increment, geometric, 4)
        
    def setTriangleTableContent(self):
        point_val = self.spin_initialPayoff.value()
        second_val = self.spin_delay_triangle.value()
        for i in range(6):
            new_item = QTableWidgetItem(str(point_val))
            self.table_Triangle.setItem(1,i,new_item)
            
            new_secitem = QTableWidgetItem(str(second_val))
            self.table_Triangle.setItem(0,i,new_secitem)
            
            if self.chk_geoPayoffIncre.isChecked():
                point_val = point_val * 2
            else:
                point_val = point_val + self.spin_incrementalPayoff.value()
            if self.chk_geoDelayIncre.isChecked():
                second_val = second_val * 2
            else:
                second_val = second_val + self.spin_delayincre_triangle.value()
            
    def setStarTableContent(self):
        point_val = self.spin_initialPayoff_star.value()
        second_val = self.spin_delay_star.value()
        for i in range(6):
            new_item = QTableWidgetItem(str(point_val))
            self.table_star.setItem(1,i,new_item)
            
            new_secitem = QTableWidgetItem(str(second_val))
            self.table_star.setItem(0,i,new_secitem)

            if self.chk_geoPayoffIncre_star.isChecked():
                point_val = point_val * 2
            else:
                point_val = point_val + self.spin_incrementalPayoff_star.value()
            
            if self.chk_geoDelayIncre_star.isChecked():
                second_val = second_val * 2
            else:
                second_val = second_val + self.spin_delayincre_star.value()

    def setConfig(self):
        self.close()
    def open_directory_dialog(self):
        directory = QFileDialog.getExistingDirectory(
            None, "Select a directory", "",
            QFileDialog.Option.ShowDirsOnly  | QFileDialog.Option.DontResolveSymlinks
        )
        self.txt_directory.setText(directory)

app = QApplication(sys.argv)
app.setFont(QFont("Arial", 13, -1, False))
app.setStyle('Fusion')
app_icon = QIcon(':icon.png')
app.setWindowIcon(app_icon)
splash = QSplashScreen()
splash.setFont(QFont("Arial", 13, -1, False))
splash.setPixmap(QPixmap(':splash.png'))
splash.showMessage('Loading...', Qt.AlignmentFlag.AlignBottom | Qt.AlignmentFlag.AlignCenter, QColor(Qt.GlobalColor.white))
splash.show()
time.sleep(2)
window = MainWindow()
window.setMinimumHeight(250)
window.show()
splash.finish(window)
sys.exit(app.exec())
